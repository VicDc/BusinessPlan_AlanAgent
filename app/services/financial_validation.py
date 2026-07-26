"""
Verifica aritmetica deterministica dei numeri finanziari prodotti dagli agenti.
NESSUNA chiamata LLM: solo ricalcolo Python delle relazioni tra i valori
(margine, break-even, capitale, copertura, ricavi vs mercato) e confronto con
quanto dichiarato dagli agenti, entro una tolleranza relativa.

È uno step separato DOPO la pipeline — non entra nel revision loop
dell'Orchestrator. Un check che non può essere calcolato (dati mancanti) viene
marcato 'non_verificabile', mai sollevata un'eccezione.
"""
from __future__ import annotations

import re
from pathlib import Path

DEFAULT_TOLERANCE = 0.05


def _num(value):
    """Coerce a float. None se non interpretabile.
    ponytail: parser tollerante ma naive — rimuove separatori migliaia E
    decimali; valori con decimali reali (rari nei totali di mercato/ricavi)
    vanno persi. Gestisce suffissi k/m/mln. Sufficiente per confronti su
    ricavi/costi interi."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    s = value.strip().lower()
    if not s:
        return None
    mult = 1.0
    for suf, m in (("mln", 1e6), ("mil", 1e6), ("m", 1e6), ("k", 1e3)):
        if s.endswith(suf):
            mult = m
            s = s[: -len(suf)]
            break
    digits = re.sub(r"[^0-9]", "", s)
    if not digits:
        return None
    return float(digits) * mult


def _sum_amounts(items, *keys):
    """Somma gli importi di una lista di voci [{item, amount}, ...].
    Ritorna 0.0 per lista vuota, None se non ci sono importi interpretabili
    (così il check a valle diventa non_verificabile)."""
    keys = keys or ("amount",)
    if not isinstance(items, list):
        return None
    if not items:
        return 0.0
    total = 0.0
    found = False
    for it in items:
        if not isinstance(it, dict):
            continue
        for k in keys:
            if k in it:
                n = _num(it[k])
                if n is not None:
                    total += n
                    found = True
                break
    return total if found else None


def _rel_close(a, b, tol):
    return abs(a - b) <= tol * max(abs(a), abs(b), 1e-9)


def _check(nome, dichiarato, ricalcolato, coerente, tol, nota):
    return {
        "nome": nome,
        "valore_dichiarato": dichiarato,
        "valore_ricalcolato": ricalcolato,
        "coerente": coerente,
        "tolleranza_usata": tol,
        "nota": nota,
    }


def _non_verificabile(nome, tol, nota):
    return _check(nome, None, None, "non_verificabile", tol, nota)


def validate_financials(
    financial_data: dict,
    funding_data: dict,
    market_data: dict,
    tolerance: float = DEFAULT_TOLERANCE,
) -> dict:
    """Ricalcola in modo deterministico le relazioni tra i numeri prodotti dagli
    agenti e confronta con i valori dichiarati.

    Ritorna un dict con:
    - 'checks': lista di {nome, valore_dichiarato, valore_ricalcolato,
      coerente (bool | 'non_verificabile'), tolleranza_usata, nota}
    - 'overall_coherent': bool (nessun check risulta incoerente)
    """
    financial_data = financial_data or {}
    funding_data = funding_data or {}
    market_data = market_data or {}
    tol = tolerance
    checks = []
    cost = financial_data.get("cost_breakdown", {}) or {}
    pricing = financial_data.get("pricing", {}) or {}

    # --- 1. Margine unitario = prezzo - costi variabili/unità ---
    recalc_margin = None
    try:
        unit_price = _num(pricing.get("unit_price_eur"))
        var_costs = _sum_amounts(cost.get("variable_costs_per_unit_eur", []))
        declared_margin = _num(pricing.get("unit_margin_eur"))
        if unit_price is None or var_costs is None:
            checks.append(_non_verificabile(
                "Margine unitario", tol,
                "Prezzo unitario o costi variabili per unità mancanti."))
        else:
            recalc_margin = unit_price - var_costs
            if declared_margin is None:
                checks.append(_check(
                    "Margine unitario", None, round(recalc_margin, 2),
                    "non_verificabile", tol, "Margine dichiarato assente."))
            else:
                coerente = _rel_close(declared_margin, recalc_margin, tol)
                checks.append(_check(
                    "Margine unitario", round(declared_margin, 2),
                    round(recalc_margin, 2), coerente, tol,
                    f"prezzo {unit_price} - costi variabili {var_costs}"))
    except Exception as exc:
        checks.append(_non_verificabile("Margine unitario", tol, f"Errore: {exc}"))

    # --- 2. Break-even = costi fissi / margine ricalcolato ---
    try:
        direct = _sum_amounts(cost.get("fixed_costs_eur_month", []))
        indirect = _sum_amounts(cost.get("indirect_costs_eur_month", []))
        # I costi indiretti (commercialista, utenze, ecc.) sono costi fissi a
        # tutti gli effetti: gli agenti li includono nel break-even, quindi li
        # sommiamo ai fissi diretti. Una lista mancante conta come 0.
        fixed = None if direct is None and indirect is None else (direct or 0.0) + (indirect or 0.0)
        declared_be = _num((financial_data.get("break_even", {}) or {}).get("units_per_month"))
        if fixed is None or not recalc_margin:
            checks.append(_non_verificabile(
                "Break-even (unità/mese)", tol,
                "Costi fissi mancanti o margine ricalcolato nullo/indisponibile."))
        else:
            recalc_be = fixed / recalc_margin
            if declared_be is None:
                checks.append(_check(
                    "Break-even (unità/mese)", None, round(recalc_be, 2),
                    "non_verificabile", tol, "Break-even dichiarato assente."))
            else:
                coerente = _rel_close(declared_be, recalc_be, tol)
                checks.append(_check(
                    "Break-even (unità/mese)", round(declared_be, 2),
                    round(recalc_be, 2), coerente, tol,
                    f"costi fissi diretti {direct or 0.0} + indiretti {indirect or 0.0} "
                    f"= {round(fixed, 2)} / margine {round(recalc_margin, 2)}"))
    except Exception as exc:
        checks.append(_non_verificabile("Break-even (unità/mese)", tol, f"Errore: {exc}"))

    # --- 3. Somma voci capitale iniziale vs eventuale totale dichiarato ---
    total_need = None
    try:
        total_need = _sum_amounts(
            financial_data.get("initial_capital", []), "amount_eur", "amount")
        declared_total = _num(financial_data.get("initial_capital_total_eur"))
        if total_need is None:
            checks.append(_non_verificabile(
                "Somma capitale iniziale", tol, "Voci capitale iniziale mancanti."))
        elif declared_total is None:
            checks.append(_check(
                "Somma capitale iniziale", None, round(total_need, 2),
                "non_verificabile", tol,
                "Nessun totale dichiarato con cui confrontare la somma."))
        else:
            coerente = _rel_close(declared_total, total_need, tol)
            checks.append(_check(
                "Somma capitale iniziale", round(declared_total, 2),
                round(total_need, 2), coerente, tol, "Somma voci vs totale dichiarato."))
    except Exception as exc:
        checks.append(_non_verificabile("Somma capitale iniziale", tol, f"Errore: {exc}"))

    # --- 4. Copertura = capitale proprio / fabbisogno (regola 25-30%) ---
    try:
        available = _num(funding_data.get("available_capital_eur"))
        declared_meets = (funding_data.get("own_capital_check", {}) or {}).get("meets_25_30_rule")
        # Fabbisogno: totale dichiarato se presente, altrimenti somma delle voci.
        declared_total = _num(financial_data.get("initial_capital_total_eur"))
        fabbisogno = declared_total if declared_total else total_need
        if available is None or not fabbisogno:
            checks.append(_non_verificabile(
                "Copertura capitale proprio", tol,
                "Capitale proprio o fabbisogno totale non disponibili."))
        else:
            coverage = available / fabbisogno
            in_band = 0.25 <= coverage <= 0.30
            pct = round(coverage * 100, 1)
            nota = (f"copertura {coverage:.1%} su fabbisogno {round(fabbisogno, 2)} "
                    f"({'rientra' if in_band else 'fuori'} banda 25-30%)")
            if not isinstance(declared_meets, bool):
                checks.append(_check(
                    "Copertura capitale proprio", declared_meets, pct,
                    "non_verificabile", tol, nota + "; own_capital_check assente."))
            else:
                # Coerente se la valutazione del modulo (dentro/fuori banda)
                # coincide con meets_25_30_rule dichiarato dall'agente.
                coerente = (in_band == declared_meets)
                checks.append(_check(
                    "Copertura capitale proprio", declared_meets, pct,
                    coerente, tol, nota))
    except Exception as exc:
        checks.append(_non_verificabile("Copertura capitale proprio", tol, f"Errore: {exc}"))

    # --- 5. Ricavo anno 1 (scenario base) non deve superare il SOM ---
    try:
        scenarios = financial_data.get("scenarios", []) or []
        base = None
        for sc in scenarios:
            if isinstance(sc, dict) and "base" in str(sc.get("scenario", "")).lower():
                base = sc
                break
        if base is None and scenarios and isinstance(scenarios[0], dict):
            base = scenarios[0]
        year1 = _num(base.get("year1_revenue_eur")) if base else None
        som = _num((market_data.get("market_sizing", {}) or {}).get("som_eur"))
        if year1 is None or som is None:
            checks.append(_non_verificabile(
                "Ricavo anno 1 vs SOM", tol,
                "Ricavo anno 1 (scenario base) o SOM di mercato non disponibili."))
        else:
            coerente = year1 <= som * (1 + tol)
            checks.append(_check(
                "Ricavo anno 1 vs SOM", round(year1, 2), round(som, 2), coerente, tol,
                "Il ricavo anno 1 non deve superare il SOM."))
    except Exception as exc:
        checks.append(_non_verificabile("Ricavo anno 1 vs SOM", tol, f"Errore: {exc}"))

    overall = not any(c["coerente"] is False for c in checks)
    return {"checks": checks, "overall_coherent": overall}


def _coerente_str(v) -> str:
    return {True: "Sì", False: "No"}.get(v, str(v))


def validation_to_markdown(result: dict) -> str:
    """Sezione markdown con la tabella dei check — stessa struttura dell'xlsx."""
    lines = [
        "## Verifica Aritmetica dei Dati Finanziari",
        "",
        f"**Coerenza complessiva:** {'Sì' if result.get('overall_coherent') else 'No'}",
        "",
        "| Controllo | Dichiarato | Ricalcolato | Coerente | Nota |",
        "| --- | --- | --- | --- | --- |",
    ]
    for c in result.get("checks", []):
        lines.append(
            f"| {c['nome']} | {c['valore_dichiarato']} | {c['valore_ricalcolato']} "
            f"| {_coerente_str(c['coerente'])} | {c['nota']} |"
        )
    return "\n".join(lines)


def export_validation_xlsx(validation_result: dict, output_path: str) -> str:
    """Scrive un .xlsx (openpyxl): una riga per check, colonne
    (Controllo, Dichiarato, Ricalcolato, Coerente, Nota). Righe non coerenti
    evidenziate in rosso. Ritorna il path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Verifica Finanziaria"
    ws.append(["Controllo", "Dichiarato", "Ricalcolato", "Coerente", "Nota"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for chk in validation_result.get("checks", []):
        coerente = chk.get("coerente")
        ws.append([
            chk.get("nome", ""),
            chk.get("valore_dichiarato"),
            chk.get("valore_ricalcolato"),
            _coerente_str(coerente),
            chk.get("nota", ""),
        ])
        if coerente is False:
            for cell in ws[ws.max_row]:
                cell.fill = red

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)
